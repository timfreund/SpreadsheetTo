import argparse
import csv
import mimetypes
import os
import openpyxl
import sys
import xlrd

implementation_names = [
    'XlsSpreadsheet',
    'XlsxSpreadsheet',
]

def open_spreadsheet(filename):
    mimeguess = mimetypes.guess_type(filename)[0]
    for impl_name in implementation_names:
        impl = getattr(sys.modules[__name__], impl_name)
        if impl.mimetype == mimeguess:
            return impl(filename)
    raise Exception("No implementation available for %s files" % mimeguess)

class Spreadsheet():
    mimetype = None

    def __init__(self, filename):
        self.filename = filename
        self.worksheets = []
        self.sheet_map = {}
        self._current_sheet = 0

    def __iter__(self):
        return self

    def __getitem__(self, key):
        if isinstance(key, int):
            return self.worksheets[key]
        return self.sheet_map(key)

    def __next__(self):
        i = self._current_sheet

        if len(self.worksheets) <= i:
            self._current_sheet = 0
            raise StopIteration

        self._current_sheet = i + 1
        return self.worksheets[i]

class Worksheet():
    def __init__(self, name=None):
        self.name = name
        self._current_row = 1

    def __getitem__(self, key):
        raise RuntimeError("Not implemented")

    def __iter__(self):
        return self

    def __next__(self):
        i = self._current_row
        if self.get_row_count() < i:
            self._current_row = 1
            raise StopIteration
        self._current_row = i + 1
        return self.get_row(i)

    def get_column(self, idx):
        raise RuntimeError("Not implemented")

    def get_column_count(self):
        raise RuntimeError("Not implemented")

    def get_name(self):
        return self.name

    def get_row(self, idx):
        raise RuntimeError("Not implemented")

    def get_row_count(self):
        raise RuntimeError("Not implemented")


class XlsSpreadsheet(Spreadsheet):
    mimetype = 'application/vnd.ms-excel'

    def __init__(self, *args):
        Spreadsheet.__init__(self, *args)
        self.workbook = xlrd.open_workbook(self.filename)
        for ws in self.workbook.sheets():
            worksheet = XlsWorksheet(ws, ws.name)
            self.worksheets.append(worksheet)
            self.sheet_map[ws.name] = worksheet

class XlsWorksheet(Worksheet):
    def __init__(self, xlrd_worksheet, name=None):
        Worksheet.__init__(self)
        self.worksheet = xlrd_worksheet
        self.name = name

    def get_column(self, idx):
        return self.worksheet.col(idx)

    def get_column_count(self):
        return self.worksheet.ncols

    def get_row_count(self):
        return self.worksheet.nrows

    def get_row(self, idx):
        # We're using a 1 based index because that's what Excel and other
        # spreadsheets use, but xlrd uses 0 based indexing, thus the -1 below
        if idx == 0:
            raise RuntimeError('Spreadsheet rows start at 1')
        xlr_row = self.worksheet.row(idx - 1)
        row = []
        for cell in xlr_row:
            row.append(cell.value)
        return row

class XlsxSpreadsheet(Spreadsheet):
    mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def __init__(self, *args):
        Spreadsheet.__init__(self, *args)
        self.workbook = openpyxl.load_workbook(self.filename, read_only=True)
        for ws in self.workbook.worksheets:
            worksheet = XlsxWorksheet(ws, ws.title)
            self.worksheets.append(worksheet)
            self.sheet_map[ws.title] = worksheet

class XlsxWorksheet(Worksheet):
    def __init__(self, openpyxl_worksheet, name=None):
        Worksheet.__init__(self)
        self.worksheet = openpyxl_worksheet
        self.name = name
        self.rows = None

    def get_column_count(self):
        return self.worksheet.max_column

    def get_row_count(self):
        return self.worksheet.max_row

    def get_row(self, idx):
        if idx == 0:
            raise RuntimeError('Spreadsheet rows start at 1')

        if not self.rows:
            self.preload_rows()

        return self.rows[idx]

    def preload_rows(self):
        self.rows = [None,]
        for row_impl in self.worksheet.rows:
            row = []
            for cell in row_impl:
                row.append(cell.value)
            self.rows.append(row)

def cli():
    parser = argparse.ArgumentParser(description="Convert spreadsheets to other formats")
    parser.add_argument('--source', dest='source',
                        required=True,
                        help='the spreadsheet we will convert')
    parser.add_argument('--destination', dest='destination',
                        default=None,
                        help='the directory where output will go (defaults to source directory)')

    args = parser.parse_args()
    workbook = open_spreadsheet(args.source)
    dest_dir = args.destination
    if not dest_dir:
        dest_dir = os.path.dirname(args.source)
        if not len(dest_dir):
            dest_dir = '.'
    base_name = os.path.sep.join((dest_dir, os.path.basename(args.source).rsplit('.', 1)[0]))

    for sheet in workbook:
        with open("%s-%s.csv" % (base_name, sheet.name), 'w') as output_file:
            writer = csv.writer(output_file)
            for row in sheet:
                writer.writerow(row)

    # import IPython
    # IPython.embed()
